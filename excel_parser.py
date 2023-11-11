from openpyxl import Workbook
from logger import logger


class ExcelParser:
    def __init__(self, book: Workbook, db_api):
        self.db_api = db_api
        self.point_sheet = book.worksheets[book.sheetnames.index('Входные данные для анализа')]
        self.employee_sheet = book.worksheets[book.sheetnames.index('Справочник сотрудников')]

    @staticmethod
    def _validate(data: dict):
        if None in data.values():
            return False
        else:
            return True

    def parse_employees(self):
        for index_row in range(2, self.employee_sheet.max_row + 1):
            employee = {
                'full_name': self.employee_sheet[index_row][0].value,
                'location': self.employee_sheet[index_row][1].value,
                'grade': self.employee_sheet[index_row][2].value
            }
            location = {'address': employee['location']}
            if self._validate(employee):
                self.db_api.try_to_set_table_row(self.db_api.Location, **location)
                employee['location'] = self.db_api.Location.get(self.db_api.Location.address == employee['location'])
                self.db_api.try_to_set_table_row(self.db_api.Employee, **employee)
        logger.info(f'Загружены данные о сотрудниках')

    def parse_points(self):
        for index_row in range(2, self.point_sheet.max_row + 1):
            point = {
                'location': self.point_sheet[index_row][1].value,
                'is_connected_yesterday': True if self.point_sheet[index_row][2].value == 'вчера' else False,
                'have_cards_and_materials': True if self.point_sheet[index_row][3].value == 'да' else False,
                'days_from_last_card': self.point_sheet[index_row][4].value,
                'accepted_requests': self.point_sheet[index_row][5].value,
                'issued_cards': self.point_sheet[index_row][6].value
            }
            location = {'address': point['location']}
            if self._validate(point):
                self.db_api.try_to_set_table_row(self.db_api.Location, **location)
                point['location'] = self.db_api.Location.get(self.db_api.Location.address == point['location'])
                self.db_api.try_to_set_table_row(self.db_api.Point, **point)
        logger.info(f'Загружены данные о точках')
